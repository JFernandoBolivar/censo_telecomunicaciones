import json
from datetime import date
from django.core.management.base import BaseCommand
from apps.users.models import Usuario
from apps.encuestas.models.models_encuestas import (
    Configuracion,
    Seccion,
    TipoPregunta,
    PreguntaPersonalizada,
    PreguntaOpciones,
)
import openpyxl


class Command(BaseCommand):
    help = "Puebla la base de datos censo desde el archivo Excel de la encuesta"

    def handle(self, *args, **options):
        excel_path = "/home/hmachado/Documentos/censo_telecomunicaciones/Encuesta Infraestructura-VGT_29062026.xlsx"
        
        self.stdout.write("Borrando datos anteriores de la encuesta...")
        PreguntaOpciones.objects.all().delete()
        PreguntaPersonalizada.objects.all().delete()
        Configuracion.objects.all().delete()
        Seccion.objects.all().delete()
        TipoPregunta.objects.all().delete()

        # 1. Crear Admin por defecto (user_creator)
        admin_email = "admin@conatel.gob.ve"
        admin_user, created = Usuario.objects.get_or_create(
            email=admin_email,
            defaults={"nombre_empresa": "Super Admin", "is_staff": True}
        )
        if created:
            admin_user.set_password("admin123")
            admin_user.save()

        # 2. Crear Configuración Inicial
        config = Configuracion.objects.create(
            fecha_inicio=date.today(),
            obligatorio=True,
            user_creator=admin_user,
        )

        # 3. Crear Secciones
        sec_info = Seccion.objects.create(seccion="INFORMACIÓN INSTITUCIONAL")
        sec_desp = Seccion.objects.create(seccion="DESPLIEGUE Y COMPARTICIÓN DE INFRAESTRUCTURA")
        sec_vgt = Seccion.objects.create(seccion="VIAS GENERALES DE TELECOMUNICACIONES (VGT) Y COMPARTICIÓN DE INFRAESTRUCTURA")

        # 4. Crear Tipos de Pregunta
        tipo_texto_corto = TipoPregunta.objects.create(nombre="Texto Corto")
        tipo_texto_largo = TipoPregunta.objects.create(nombre="Texto Largo")
        tipo_entero = TipoPregunta.objects.create(nombre="Numérico Entero")
        tipo_decimal = TipoPregunta.objects.create(nombre="Numérico Decimal")
        tipo_sino = TipoPregunta.objects.create(nombre="Selección Única (SI/NO)")
        tipo_multiple = TipoPregunta.objects.create(nombre="Selección Múltiple")
        tipo_geografica = TipoPregunta.objects.create(nombre="Selección Geográfica")

        # 5. Cargar Excel
        self.stdout.write(f"Cargando Excel desde: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Omitir filas redundantes y sub-preguntas convertidas en opciones
        omitir_filas_b = ["1.1", "1.4", "3.1.1.1", "3.1.1.2", "3.1.1.3", "3.1.1.4"]

        for row_idx in range(7, 65):
            col_b = sheet.cell(row=row_idx, column=2).value
            col_c = sheet.cell(row=row_idx, column=3).value
            col_f = sheet.cell(row=row_idx, column=6).value
            col_j = sheet.cell(row=row_idx, column=10).value

            if not col_c:
                continue

            # Determinar Sección actual
            seccion_actual = None
            str_b = str(col_b).strip() if col_b else ""
            if str_b.startswith("1."):
                seccion_actual = sec_info
            elif str_b.startswith("2."):
                seccion_actual = sec_desp
            elif str_b.startswith("3."):
                seccion_actual = sec_vgt

            # Omitir si es una cabecera principal ("1", "2", "3") o redundante
            if str_b in ["1", "2", "3"] or str_b in omitir_filas_b:
                continue

            # Solo procesar preguntas reales que pertenezcan a una sección
            if not seccion_actual:
                continue

            titulo = col_c
            texto_f = str(col_f).lower() if col_f else ""
            texto_j = str(col_j).lower() if col_j else ""

            # Determinar Tipo y Validación
            tipo_asignado = tipo_texto_corto
            validacion = {
                "regex": "",
                "min_chars_or_digits": 0,
                "max_chars_or_digits": 255
            }
            opciones_sino = False

            if "si/no" in texto_j or "si o no" in texto_f:
                tipo_asignado = tipo_sino
                opciones_sino = True
                validacion["max_chars_or_digits"] = 2
                validacion["regex"] = "^(SI|NO)$"
            elif "numérica" in texto_j and "decimales" in texto_j:
                tipo_asignado = tipo_decimal
                validacion["regex"] = "^\\d+(\\.\\d{1,2})?$"
            elif "numérica" in texto_j and "enteros" in texto_j:
                tipo_asignado = tipo_entero
                validacion["regex"] = "^\\d+$"
            elif "parroquias" in titulo.lower() or "municipios" in titulo.lower() or "entidades" in titulo.lower():
                tipo_asignado = tipo_geografica
                validacion["max_chars_or_digits"] = 1000
                validacion["regex"] = "^\\d+(,\\d+)*$"
            elif "informante" in titulo.lower() or "cargo" in titulo.lower():
                tipo_asignado = tipo_texto_corto
                validacion["max_chars_or_digits"] = 255
                validacion["regex"] = "^[a-zA-ZáéíóúÁÉÍÓÚñÑ\\s]+$"
            elif "detalle" in titulo.lower() or "especifique" in titulo.lower() or "sugerencias" in titulo.lower() or "opiniones" in texto_j or "ductos" in titulo.lower() or "postes" in titulo.lower() or "otros" in titulo.lower():
                tipo_asignado = tipo_texto_largo if "detalle" in titulo.lower() or "sugerencias" in titulo.lower() else tipo_texto_corto
                validacion["max_chars_or_digits"] = 2000 if tipo_asignado == tipo_texto_largo else 255
                validacion["regex"] = "^[\\w\\s\\.,\\-ñÑáéíóúÁÉÍÓÚ]+$"
            else:
                validacion["regex"] = "^[\\w\\s\\.,\\-ñÑáéíóúÁÉÍÓÚ]+$"

            # Crear pregunta
            pregunta = PreguntaPersonalizada.objects.create(
                titulo=titulo,
                tipo_pregunta=tipo_asignado,
                configuracion=config,
                seccion=seccion_actual,
                es_obligatorio=True,
                validacion=validacion,
            )

            # Crear opciones si es SI/NO
            if opciones_sino:
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="SI")
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="NO")
                
            # Si es la 3.1.1, forzamos a múltiple y agregamos las opciones manualmente
            if str_b == "3.1.1":
                pregunta.tipo_pregunta = tipo_multiple
                pregunta.save()
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="Ductos")
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="Postes")
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="Torres")
                PreguntaOpciones.objects.create(pregunta=pregunta, opcion="Otros")

        self.stdout.write(self.style.SUCCESS("Migración de Excel a BD finalizada con éxito!"))
