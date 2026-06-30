import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.encuestas.models.models_encuestas import (
    PreguntaPersonalizada,
    UsuarioPermiso,
)


def generar_excel_respuestas(configuracion) -> io.BytesIO:
    preguntas = (
        PreguntaPersonalizada.objects
        .filter(configuracion=configuracion)
        .select_related("tipo_pregunta", "seccion")
        .prefetch_related("opciones")
        .order_by("seccion__id", "id")
    )
    pregunta_ids = list(preguntas.values_list("id", flat=True))

    if not pregunta_ids:
        output = io.BytesIO()
        wb = Workbook()
        wb.save(output)
        output.seek(0)
        return output

    permisos = (
        UsuarioPermiso.objects
        .filter(pregunta_personalizada__in=pregunta_ids)
        .select_related("usuario")
        .order_by("usuario__email")
    )

    usuarios_unicos = OrderedDictUnique()
    for permiso in permisos:
        usuarios_unicos[permiso.usuario_id] = permiso.usuario

    respuesta_encuesta_map = {}
    respuesta_opciones_map = defaultdict(list)

    for permiso in permisos:
        for re in permiso.respuestaencuesta_set.all():
            respuesta_encuesta_map[(permiso.usuario_id, permiso.pregunta_personalizada_id)] = re.respuesta
        for ro in permiso.respuestaopciones_set.select_related("pregunta_opciones"):
            respuesta_opciones_map[(permiso.usuario_id, permiso.pregunta_personalizada_id)].append(
                ro.pregunta_opciones.opcion
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Encuesta"

    headers = ["#", "RIF", "Empresa (Nombre)", "Email"]
    for p in preguntas:
        headers.append(p.titulo)

    ws.append(headers)

    for idx, user in enumerate(usuarios_unicos.values(), start=1):
        row = [idx, user.rif_empresa or "", user.nombre_empresa, user.email]
        for pid in pregunta_ids:
            texto = respuesta_encuesta_map.get((user.id, pid), "")
            if texto:
                row.append(texto)
            else:
                opciones_lista = respuesta_opciones_map.get((user.id, pid), [])
                row.append("; ".join(opciones_lista) if opciones_lista else "")
        ws.append(row)

    _aplicar_estilos(ws, len(headers))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class OrderedDictUnique(dict):
    def __setitem__(self, key, value):
        if key not in self:
            super().__setitem__(key, value)


def _aplicar_estilos(ws, num_columnas):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    data_alignment = Alignment(vertical="center")

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_col=num_columnas), start=2):
        for cell in row_cells:
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_idx % 2 == 0:
                cell.fill = even_fill

    CHAR_WIDTH = 1.2
    MIN_WIDTH = 12
    MAX_WIDTH = 55
    for col_idx in range(1, num_columnas + 1):
        col_letter = get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value
        header_len = len(str(header_val)) if header_val else 0
        max_data_len = 0
        for row_cells in ws.iter_rows(min_row=2, max_col=col_idx, min_col=col_idx, values_only=True):
            val = row_cells[0]
            val_str = str(val) if val is not None else ""
            max_data_len = max(max_data_len, len(val_str))
        best_len = max(header_len, max_data_len)
        col_width = min(max(best_len * CHAR_WIDTH + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[col_letter].width = col_width

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
